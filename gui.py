from calendar import day_name
from tkcalendar import DateEntry
from tkinter import *
from tkinter import ttk, messagebox
from PIL import ImageTk, Image
from datetime import datetime
from ttkwidgets.autocomplete import AutocompleteEntry, AutocompleteCombobox
from win32com.client import Dispatch
from pywintypes import com_error
import openpyxl, subprocess, win32com.client, re, os
import pandas as pd
import sys, os


pathfiles = os.getcwd()
#SET THE XLSX FILE
# print(bool(os.getcwd()+'\\ne_pas_toucher.xlsx'))
xfile = openpyxl.load_workbook(os.getcwd()+'\\ne_pas_toucher.xlsx')
sheet = xfile.get_sheet_by_name(xfile.sheetnames[0])
sheetname_bdc2 = xfile.sheetnames

file = pd.ExcelFile(os.getcwd()+'\\BDC - EQUO CONSTRUCTION.xlsx')
sheetnames_bdc = file.sheet_names




nom_fournisseur, adr_fournisseur, cp_fournisseur, contact_fournisseur, numero_fournisseur = [], [], [], [], []
chantier, adr_chantier, cp_chantier, contact_chantier, numero_chantier = [], [], [], [], []

fournisseur = {"nom" : nom_fournisseur, 
               "adresse" : adr_fournisseur, 
               "cp" : cp_fournisseur, 
               "contact" : contact_fournisseur, 
               "numero" : numero_fournisseur}

livraison = {"chantier" : chantier,
             "adresse" : adr_chantier, 
             "cp" : cp_chantier, 
             "contact" : contact_chantier, 
             "numero" : numero_chantier}


def set_lists():
    #TXT FILES FOURNISSEUR
    with open("Fentreprise.txt", "r", encoding ='utf-8') as file:
        content = file.read().splitlines()
        for item in content:
            fournisseur["nom"].append(item)

    with open("Fadresse.txt", "r", encoding ='utf-8') as file:
        content = file.read().splitlines()
        for item in content:
            fournisseur["adresse"].append(item)

    with open("Fville.txt", "r", encoding ='utf-8') as file:
        content = file.read().splitlines()
        for item in content:
            fournisseur["cp"].append(item)

    #TXT FILES LIVRAISON
    with open("Lchantier.txt", "r", encoding ='utf-8') as file:
        content = file.read().splitlines()
        for item in content:
            livraison["chantier"].append(item)

    with open("Ladresse.txt", "r", encoding ='utf-8') as file:
        content = file.read().splitlines()
        for item in content:
            livraison["adresse"].append(item)

    with open("Lville.txt", "r", encoding ='utf-8') as file:
        content = file.read().splitlines()
        for item in content:
            livraison["cp"].append(item)
    
    #TXT FILES CONTACTS
    fiche_contact_L, fiche_contact_F = [], []
    with open("Lcontact.txt", "r", encoding ='utf-8') as file:
        content = file.read().splitlines()
        for item in content:
            livraison["contact"].append(item)

    with open("Lnumero.txt", "r", encoding ='utf-8') as file:
        content = file.read().splitlines()
        for item in content:
            livraison["numero"].append(item)

    with open("Fcontact.txt", "r", encoding ='utf-8') as file:
        content = file.read().splitlines()
        for item in content:
            fournisseur["contact"].append(item)
    
    with open("Fnumero.txt", "r", encoding ='utf-8') as file:
        content = file.read().splitlines()
        for item in content:
            fournisseur["numero"].append(item)
    # print(len(fournisseur["contact"]))
    # print(len(fournisseur["numero"]))
    for i in range(len(fournisseur["contact"])):
        fiche_contact_F.append(str(fournisseur["contact"][i])+" : "+str(fournisseur["numero"][i]))
    print(len(livraison["contact"]))
    print(len(livraison["numero"]))
    for i in range(len(livraison["contact"])):
        fiche_contact_L.append(str(livraison["contact"][i])+" : "+str(livraison["numero"][i]))
    
    # print(fiche_contact_L[0].rsplit('\n', 1))
    return fiche_contact_F, fiche_contact_L

facturation = {"departement" : "Service Comptabilité", 
                "entreprise" : "EQUO CONSTRUCTION", 
                "adresse" : "16 Rue Ampère", 
                "cp" : "95300 - Pontoise"}


fiche_contact_F, fiche_contact_L = set_lists()
# print(fiche_contact_L)



now = datetime.now()
date = now.strftime("%d/%m/%Y")
import datetime
now2 = now + datetime.timedelta(days = 3)
date_livraison = now2.strftime("%d/%m/%Y")
prenom, nom = "Quentin", "Gourier"
redac = prenom + nom
reference = ""

informations = {"date" : date, 
                "rédacteur" : redac, 
                "num_commande" : prenom[0]+nom[0]+"-",
                "date de livraison" : date_livraison, 
                "référence chantier" : reference}


#TK
window = Tk()
window.resizable(False, False)
window.title('MyBDC')
font0 = ('Consolas', 25)
font1 = ('Consolas', 20)
font2 = ('Consolas', 16)
font3 = ('Consolas', 13)
font4 = ('Consolas', 12)
font4bis = ('Consolas italic', 12)
font5 = ('Consolas Bold', 14)
font6 = ('Consolas', 11)
font7 = ('Consolas Bold', 13)

win1 = PanedWindow(window)
win1.grid(row = 0, rowspan = 2, column = 1)
# win1.config(bg = 'grey59')
win2 = PanedWindow(window)
win2.grid(row = 0, rowspan = 3, column = 2)
# win1 = PanedWindow(window)
# win1.grid(row = 3, column = 1)

titre = Label(win1, text =  " - BON DE COMMANDE - ")
titre.grid(row = 0, column = 1, columnspan = 7, sticky = W+E)
titre.config(font = font0, borderwidth=30)

blank = Label(win1, text = "", borderwidth=20)
blank.grid(row = 0, rowspan = 12, column = 0)

blank2 = Label(win1, text = "")
blank2.grid(row = 0, rowspan = 12, column = 7, sticky = N+S)

blank3 = Label(win1, text = "")
blank3.grid(row = 11, column = 0, columnspan = 7)

# blank4 = Label(win1, text = "")
# blank4.grid(row = 6, column = 0, columnspan = 7)

blank = Label(win1, text = "", borderwidth=20)
blank.grid(row = 0, rowspan = 12, column = 8)

#LIVRAISON
l_livraison = Label(win1, text = "Livraison")
l_livraison.grid(row = 1, column = 1, columnspan = 2, sticky = W)
l_livraison.config(font = font2, borderwidth=20)

def auto_livraison():
    # print(e_chantier.get())
    e_adresse.delete(0, END)
    e_cp.delete(0, END)
    for i in range(len(livraison["chantier"])):
        if e_chantier.get() == livraison["chantier"][i]:
            # print(i)
            e_adresse.insert(0,livraison["adresse"][i])
            e_cp.insert(0,livraison["cp"][i])
open_file = Image.open(r"autofill.png")
file = open_file.resize((30,30), Image.ANTIALIAS)
fleche1 = ImageTk.PhotoImage(file)
autofill_livraison = Button(win1, text = "   auto", command = auto_livraison, image = fleche1, borderwidth=10)
autofill_livraison.grid(row = 3, rowspan = 2, column = 0, sticky = N+S)
autofill_livraison.config(font = font6, relief = FLAT)

def handle_chantier(_):
    e_chantier.delete(0, END)
def handle_adresse(_):
    e_adresse.delete(0, END)
def handle_cp(_):
    e_cp.delete(0, END)
def handle_contact(_):
    e_contact.delete(0, END)


e_chantier = AutocompleteEntry(win1, width = 32, completevalues = livraison["chantier"])
e_chantier.grid(row = 2, column = 1, columnspan = 2, sticky = W)
e_chantier.config(font = font6)
e_chantier.insert(0, "nom du chantier")
e_chantier.bind("<FocusIn>", handle_chantier)

e_adresse = AutocompleteEntry(win1, width = 32, completevalues = livraison["adresse"])
e_adresse.grid(row = 3, column = 1, columnspan = 2, sticky = W)
e_adresse.config(font = font6)
e_adresse.insert(0, "adresse")
e_adresse.bind("<FocusIn>", handle_adresse)

e_cp = AutocompleteEntry(win1, width = 32, completevalues = livraison["cp"])
e_cp.grid(row = 4, column = 1, columnspan = 2, sticky = W)
e_cp.config(font = font6)
e_cp.insert(0, "ville")
e_cp.bind("<FocusIn>", handle_cp)

e_contact= AutocompleteEntry(win1, width = 32, completevalues = fiche_contact_L)
e_contact.grid(row = 5, column = 1, columnspan = 2, sticky = W)
e_contact.config(font = font6)
e_contact.insert(0, "contact")
e_contact.bind("<FocusIn>", handle_contact)


#FOURNISSEUR
l_fournisseur = Label(win1, text = "    Fournisseur")
l_fournisseur.grid(row = 1, column = 2)
l_fournisseur.config(font = font2, borderwidth=20)

def auto_fournisseur():
    # print(e_nom.get())
    e_adresse2.delete(0, END)
    e_cp2.delete(0, END)
    e_contact2.delete(0, END)
    for i in range(len(fournisseur["nom"])):
        if e_nom.get() == fournisseur["nom"][i]:
            # print(i)
            e_adresse2.insert(0,fournisseur["adresse"][i])
            e_cp2.insert(0,fournisseur["cp"][i])
            e_contact2.insert(0, fiche_contact_F[i])

open_file = Image.open(r"autofill.png")
file = open_file.resize((30,30), Image.ANTIALIAS)
fleche2 = ImageTk.PhotoImage(file)
autofill_fournisseur = Button(win1, text = "auto", command = auto_fournisseur, image = fleche2, borderwidth=10)
autofill_fournisseur.grid(row = 3, rowspan = 3, column = 3, sticky = N+S)
autofill_fournisseur.config(font = font6, relief = FLAT)

def handle_nom(_):
    e_nom.delete(0, END)
def handle_adresse2(_):
    e_adresse2.delete(0, END)
def handle_cp2(_):
    e_cp2.delete(0, END)
def handle_contact2(_):
    e_contact2.delete(0, END)

def handle_devis(_):
    numero_devis.delete(0, END)

e_nom = AutocompleteEntry(win1, width = 32, completevalues = fournisseur["nom"])
e_nom.grid(row = 2, column = 2, sticky = E)
e_nom.config(font = font6)
e_nom.insert(0, "nom du fournisseur")
e_nom.bind("<FocusIn>", handle_nom)

e_adresse2 = AutocompleteEntry(win1, width = 32, completevalues = fournisseur["adresse"])
e_adresse2.grid(row = 3, column = 2, sticky = E)
e_adresse2.config(font = font6)
e_adresse2.insert(0, "adresse")
e_adresse2.bind("<FocusIn>", handle_adresse2)

e_cp2 = AutocompleteEntry(win1, width = 32, completevalues = fournisseur["cp"])
e_cp2.grid(row = 4, column = 2, sticky = E)
e_cp2.config(font = font6)
e_cp2.insert(0, "ville")
e_cp2.bind("<FocusIn>", handle_cp2)

e_contact2 = AutocompleteEntry(win1, width = 32, completevalues = fiche_contact_F)
e_contact2.grid(row = 5, column = 2, sticky = E)
e_contact2.config(font = font6)
e_contact2.insert(0, "contact")
e_contact2.bind("<FocusIn>", handle_contact2)

#PRECISIONS
l_precisions = Label(win1, text = "Precisions")
l_precisions.grid(row = 7, column = 1, columnspan = 3, sticky = W)
l_precisions.config(font = font2, borderwidth = 20)
precisions = Text(win1, height = 4, width = 68)
precisions.grid(row = 8, column = 1, columnspan = 2, sticky = W+E)

blank111 = Label(win1, text = " ", fg = 'grey')
blank111.grid(row = 9, column = 2, sticky = E)
blank111.config(font = font4)

#INFORMATIONS
list_bat = [""]
numero_devis = AutocompleteEntry(win1, width = 32, completevalues = list_bat)
numero_devis.grid(row = 11, column = 1, sticky = W)
numero_devis.config(font = font6)
numero_devis.insert(0, "n° devis")
numero_devis.bind("<FocusIn>", handle_devis)

l_informations = Label(win1, text = "Informations")
l_informations.grid(row = 10, column = 1, columnspan = 3, sticky = W)
l_informations.config(font = font2, borderwidth=20)

salaries = ["Bussem CAMURCU", "Hassan ABLOUH", "Anis ACHOUR", "Abdelhamid ABOUD", "Mustafa DURMUS",
            "Jonathan PEREIRA", "Erhan HODZHA", "Oguz BOZKURT", "Jaime GONCALVES"]
e_redac = AutocompleteCombobox(win1, completevalues = salaries, width = 28)
e_redac.grid(row = 12, column = 1, columnspan = 2, sticky = W)
e_redac.config(width = 34)

cal = DateEntry(win1, width= 10, background= "#165108", foreground= "white",bd=2)

l_date = Label(win1, text = "Date du jour : "+date, fg = 'grey')
l_date.grid(row = 11, column = 2, sticky=E)
l_date.config(font = font4)

l_date = Label(win1, text = "Livré le :", fg = 'grey')
l_date.grid(row = 12, column = 2)
l_date.config(font = font4)

cal.grid(row = 12, column = 2, sticky=E)

blank11 = Label(win1, text = " ", fg = 'grey')
blank11.grid(row = 13, column = 2, sticky = E)
blank11.config(font = font4)


ttk.Separator(win1, orient=VERTICAL).grid(column=8, row=0, rowspan=20, sticky='ns')


#PANEL COMMANDE
def logo():
    back = pathfiles+'Logo.png'
    open_file = Image.open(back)
    file = open_file.resize((90,45), Image.ANTIALIAS)
    image = ImageTk.PhotoImage(file)

    img = Label(win2, image = image)
    img.image = image
    img.grid(row = 0, column = 1, columnspan = 2, sticky = W)
# logo()

unit = ["U", "T", "%", "ens", "jour(s)", "heure(s)"]
commande = Label(win2, text = "Commande")
commande.grid(row = 1, column = 2, columnspan = 2, sticky = W)
commande.config(font = font2, borderwidth=20)

# titre2 = Label(win2, text =  " ")
# titre2.grid(row = 0, column = 1, columnspan = 20)
# titre2.config(font = font0, borderwidth=30)

blank6 = Label(win2, text = "", borderwidth=3)
blank6.grid(row = 0, rowspan = 12, column = 1)
blank10 = Label(win2, text = "", borderwidth=20)
blank10.grid(row = 0, rowspan = 12, column = 14)


def less():
    value = int(e_q.get())
    value2 = float(e_prixunit.get())
    value -= 1
    e_q.delete(0, 'end')
    e_q.insert(0, value)
    e_prixtotal.delete(0, 'end')
    e_prixtotal.insert(0, str(float(value)*value2)+ "€")
def more():
    value = int(e_q.get())
    value2 = float(e_prixunit.get())
    value += 1
    e_q.delete(0, 'end')
    e_q.insert(0, value)
    e_prixtotal.delete(0, 'end')
    e_prixtotal.insert(0, str(float(value)*value2)+ "€")

for i in range(0,8):
    blank5 = Label(win2, text = "", borderwidth=8)
    blank5.grid(row = 2+i, column = 4)
    blank7 = Label(win2, text = "", borderwidth=8)
    blank7.grid(row = 2+i, column = 6)
    blank8 = Label(win2, text = "", borderwidth=8)
    blank8.grid(row = 2+i, column = 8)
    blank9 = Label(win2, text = "", borderwidth=8)
    blank9.grid(row = 2+i, column = 12)

e_designation = Entry(win2, width = 30, fg ="grey")
e_designation.grid(row = 2, column = 2, columnspan = 2, sticky = W)
e_designation.config(font = font4)

e_u = ttk.Combobox(win2, values = unit, width = 7)
e_u.grid(row = 2, column = 5)

e_prixunit = Entry(win2, width = 8, fg ="grey")
e_prixunit.grid(row = 2, column = 7, sticky = W)
e_prixunit.config(font = font4)

moins = Button(win2, text = "-", command = less, borderwidth=0)
moins.grid(row = 2, column = 9)
moins.config(font = font5, width = 2)
plus = Button(win2, text = "+", command = more, borderwidth=0)
plus.grid(row = 2, column = 11)
plus.config(font = font5, width = 2)

quantity = StringVar()
e_q = Entry(win2, width = 2, fg ="grey", textvariable = quantity)
e_q.grid(row = 2, column = 10, sticky = W)
e_q.config(font = font4)
e_q.insert(0, "0")

prixtotal = StringVar()
e_prixtotal = Entry(win2, width = 8, fg ="grey", textvariable = prixtotal)
e_prixtotal.grid(row = 2, column = 13, sticky = W)
e_prixtotal.config(font = font4)



 
#ith command
e_designation1 = Entry(win2, width = 30, fg ="grey")
e_designation1.grid(row = 3, column = 2, columnspan = 2, sticky = W)
e_designation1.config(font = font4)

e_u1 = ttk.Combobox(win2, values = unit, width = 7)
e_u1.grid(row = 3, column = 5)

e_prixunit1 = Entry(win2, width = 8, fg ="grey")
e_prixunit1.grid(row = 3, column = 7, sticky = W)
e_prixunit1.config(font = font4)

def less1():
    value1 = int(e_q1.get())
    value21 = float(e_prixunit1.get())
    value1 -= 1
    e_q1.delete(0, 'end')
    e_q1.insert(0, value1)
    e_prixtotal1.delete(0, 'end')
    e_prixtotal1.insert(0, str(float(value1)*value21)+ "€")
def more1():
    value1 = int(e_q1.get())
    value21 = float(e_prixunit1.get())
    value1 += 1
    e_q1.delete(0, 'end')
    e_q1.insert(0, value1)
    e_prixtotal1.delete(0, 'end')
    e_prixtotal1.insert(0, str(float(value1)*value21)+ "€")

moins1 = Button(win2, text = "-", command = less1, borderwidth=0)
moins1.grid(row = 3, column = 9)
moins1.config(font = font5, width = 2)
plus1 = Button(win2, text = "+", command = more1, borderwidth=0)
plus1.grid(row = 3, column = 11)
plus1.config(font = font5, width = 2)

quantity1 = StringVar()
e_q1 = Entry(win2, width = 2, fg ="grey", textvariable = quantity1)
e_q1.grid(row = 3, column = 10, sticky = W)
e_q1.config(font = font4)
e_q1.insert(0, "0")

prixtotal1 = StringVar()
e_prixtotal1 = Entry(win2, width = 8, fg ="grey", textvariable = prixtotal1)
e_prixtotal1.grid(row = 3, column = 13, sticky = W)
e_prixtotal1.config(font = font4)


e_designation2 = Entry(win2, width = 30, fg ="grey")
e_designation2.grid(row = 4, column = 2, columnspan = 2, sticky = W)
e_designation2.config(font = font4)

e_u2 = ttk.Combobox(win2, values = unit, width = 7)
e_u2.grid(row = 4, column = 5)

e_prixunit2 = Entry(win2, width = 8, fg ="grey")
e_prixunit2.grid(row = 4, column = 7, sticky = W)
e_prixunit2.config(font = font4)

def less2():
    value2 = int(e_q2.get())
    value22 = float(e_prixunit2.get())
    value2 -= 1
    e_q2.delete(0, 'end')
    e_q2.insert(0, value2)
    e_prixtotal2.delete(0, 'end')
    e_prixtotal2.insert(0, str(float(value2)*value22)+ "€")
def more2():
    value2 = int(e_q2.get())
    value22 = float(e_prixunit2.get())
    value2 += 1
    e_q2.delete(0, 'end')
    e_q2.insert(0, value2)
    e_prixtotal2.delete(0, 'end')
    e_prixtotal2.insert(0, str(float(value2)*value22)+ "€")

moins2 = Button(win2, text = "-", command = less2, borderwidth=0)
moins2.grid(row = 4, column = 9)
moins2.config(font = font5, width = 2)
plus2 = Button(win2, text = "+", command = more2, borderwidth=0)
plus2.grid(row = 4, column = 11)
plus2.config(font = font5, width = 2)

quantity2 = StringVar()
e_q2 = Entry(win2, width = 2, fg ="grey", textvariable = quantity2)
e_q2.grid(row = 4, column = 10, sticky = W)
e_q2.config(font = font4)
e_q2.insert(0, "0")

prixtotal2 = StringVar()
e_prixtotal2 = Entry(win2, width = 8, fg ="grey", textvariable = prixtotal2)
e_prixtotal2.grid(row = 4, column = 13, sticky = W)
e_prixtotal2.config(font = font4)



e_designation3 = Entry(win2, width = 30, fg ="grey")
e_designation3.grid(row = 5, column = 2, columnspan = 2, sticky = W)
e_designation3.config(font = font4)

e_u3 = ttk.Combobox(win2, values = unit, width = 7)
e_u3.grid(row = 5, column = 5)

e_prixunit3 = Entry(win2, width = 8, fg ="grey")
e_prixunit3.grid(row = 5, column = 7, sticky = W)
e_prixunit3.config(font = font4)

def less3():
    value3 = int(e_q3.get())
    value23 = float(e_prixunit3.get())
    value3 -= 1
    e_q3.delete(0, 'end')
    e_q3.insert(0, value3)
    e_prixtotal3.delete(0, 'end')
    e_prixtotal3.insert(0, str(float(value3)*value23)+ "€")
def more3():
    value3 = int(e_q3.get())
    value23 = float(e_prixunit3.get())
    value3 += 1
    e_q3.delete(0, 'end')
    e_q3.insert(0, value3)
    e_prixtotal3.delete(0, 'end')
    e_prixtotal3.insert(0, str(float(value3)*value23)+ "€")

moins3 = Button(win2, text = "-", command = less3, borderwidth=0)
moins3.grid(row = 5, column = 9)
moins3.config(font = font5, width = 2)
plus3 = Button(win2, text = "+", command = more3, borderwidth=0)
plus3.grid(row = 5, column = 11)
plus3.config(font = font5, width = 2)

quantity3 = StringVar()
e_q3 = Entry(win2, width = 2, fg ="grey", textvariable = quantity3)
e_q3.grid(row = 5, column = 10, sticky = W)
e_q3.config(font = font4)
e_q3.insert(0, "0")

prixtotal3 = StringVar()
e_prixtotal3 = Entry(win2, width = 8, fg ="grey", textvariable = prixtotal3)
e_prixtotal3.grid(row = 5, column = 13, sticky = W)
e_prixtotal3.config(font = font4)



e_designation4 = Entry(win2, width = 30, fg ="grey")
e_designation4.grid(row = 6, column = 2, columnspan = 2, sticky = W)
e_designation4.config(font = font4)

e_u4 = ttk.Combobox(win2, values = unit, width = 7)
e_u4.grid(row = 6, column = 5)

e_prixunit4 = Entry(win2, width = 8, fg ="grey")
e_prixunit4.grid(row = 6, column = 7, sticky = W)
e_prixunit4.config(font = font4)

def less4():
    value4 = int(e_q2.get())
    value24= float(e_prixunit2.get())
    value4 -= 1
    e_q4.delete(0, 'end')
    e_q4.insert(0, value4)
    e_prixtotal4.delete(0, 'end')
    e_prixtotal4.insert(0, str(float(value4)*value24)+ "€")
def more4():
    value4 = int(e_q4.get())
    value24 = float(e_prixunit4.get())
    value4 += 1
    e_q4.delete(0, 'end')
    e_q4.insert(0, value4)
    e_prixtotal4.delete(0, 'end')
    e_prixtotal4.insert(0, str(float(value4)*value24)+ "€")

moins4 = Button(win2, text = "-", command = less4, borderwidth=0)
moins4.grid(row = 6, column = 9)
moins4.config(font = font5, width = 2)
plus4 = Button(win2, text = "+", command = more4, borderwidth=0)
plus4.grid(row = 6, column = 11)
plus4.config(font = font5, width = 2)

quantity4 = StringVar()
e_q4 = Entry(win2, width = 2, fg ="grey", textvariable = quantity4)
e_q4.grid(row = 6, column = 10, sticky = W)
e_q4.config(font = font4)
e_q4.insert(0, "0")

prixtotal4 = StringVar()
e_prixtotal4 = Entry(win2, width = 8, fg ="grey", textvariable = prixtotal4)
e_prixtotal4.grid(row = 6, column = 13, sticky = W)
e_prixtotal4.config(font = font4)


e_designation5 = Entry(win2, width = 30, fg ="grey")
e_designation5.grid(row = 7, column = 2, columnspan = 2, sticky = W)
e_designation5.config(font = font4)

e_u5 = ttk.Combobox(win2, values = unit, width = 7)
e_u5.grid(row = 7, column = 5)

e_prixunit5 = Entry(win2, width = 8, fg ="grey")
e_prixunit5.grid(row = 7, column = 7, sticky = W)
e_prixunit5.config(font = font4)

def less5():
    value5 = int(e_q5.get())
    value25 = float(e_prixunit5.get())
    value5 -= 1
    e_q5.delete(0, 'end')
    e_q5.insert(0, value5)
    e_prixtotal5.delete(0, 'end')
    e_prixtotal5.insert(0, str(float(value5)*value25)+ "€")
def more5():
    value5 = int(e_q5.get())
    value25 = float(e_prixunit5.get())
    value5 += 1
    e_q5.delete(0, 'end')
    e_q5.insert(0, value5)
    e_prixtotal5.delete(0, 'end')
    e_prixtotal5.insert(0, str(float(value5)*value25)+ "€")

moins5 = Button(win2, text = "-", command = less5, borderwidth=0)
moins5.grid(row = 7, column = 9)
moins5.config(font = font5, width = 2)
plus5 = Button(win2, text = "+", command = more5, borderwidth=0)
plus5.grid(row = 7, column = 11)
plus5.config(font = font5, width = 2)

quantity5 = StringVar()
e_q5 = Entry(win2, width = 2, fg ="grey", textvariable = quantity5)
e_q5.grid(row = 7, column = 10, sticky = W)
e_q5.config(font = font4)
e_q5.insert(0, "0")

prixtotal5 = StringVar()
e_prixtotal5 = Entry(win2, width = 8, fg ="grey", textvariable = prixtotal5)
e_prixtotal5.grid(row = 7, column = 13, sticky = W)
e_prixtotal5.config(font = font4)



e_designation6 = Entry(win2, width = 30, fg ="grey")
e_designation6.grid(row = 8, column = 2, columnspan = 2, sticky = W)
e_designation6.config(font = font4)

e_u6 = ttk.Combobox(win2, values = unit, width = 7)
e_u6.grid(row = 8, column = 5)

e_prixunit6 = Entry(win2, width = 8, fg ="grey")
e_prixunit6.grid(row = 8, column = 7, sticky = W)
e_prixunit6.config(font = font4)

def less6():
    value6 = int(e_q6.get())
    value26 = float(e_prixunit6.get())
    value6 -= 1
    e_q6.delete(0, 'end')
    e_q6.insert(0, value6)
    e_prixtotal6.delete(0, 'end')
    e_prixtotal6.insert(0, str(float(value6)*value26)+ "€")
def more6():
    value6 = int(e_q6.get())
    value26 = float(e_prixunit6.get())
    value6 += 1
    e_q6.delete(0, 'end')
    e_q6.insert(0, value6)
    e_prixtotal6.delete(0, 'end')
    e_prixtotal6.insert(0, str(float(value6)*value26)+ "€")

moins6 = Button(win2, text = "-", command = less6, borderwidth=0)
moins6.grid(row = 8, column = 9)
moins6.config(font = font5, width = 2)
plus6 = Button(win2, text = "+", command = more6, borderwidth=0)
plus6.grid(row = 8, column = 11)
plus6.config(font = font5, width = 2)

quantity6 = StringVar()
e_q6 = Entry(win2, width = 2, fg ="grey", textvariable = quantity6)
e_q6.grid(row = 8, column = 10, sticky = W)
e_q6.config(font = font4)
e_q6.insert(0, "0")

prixtotal6 = StringVar()
e_prixtotal6 = Entry(win2, width = 8, fg ="grey", textvariable = prixtotal6)
e_prixtotal6.grid(row = 8, column = 13, sticky = W)
e_prixtotal6.config(font = font4)



e_designation7 = Entry(win2, width = 30, fg ="grey")
e_designation7.grid(row = 9, column = 2, columnspan = 2, sticky = W)
e_designation7.config(font = font4)

e_u7 = ttk.Combobox(win2, values = unit, width = 7)
e_u7.grid(row = 9, column = 5)

e_prixunit7 = Entry(win2, width = 8, fg ="grey")
e_prixunit7.grid(row = 9, column = 7, sticky = W)
e_prixunit7.config(font = font4)

def less7():
    value7 = int(e_q7.get())
    value27 = float(e_prixunit7.get())
    value7 -= 1
    e_q7.delete(0, 'end')
    e_q7.insert(0, value7)
    e_prixtotal7.delete(0, 'end')
    e_prixtotal7.insert(0, str(float(value7)*value27)+ "€")
def more7():
    value7 = int(e_q7.get())
    value27 = float(e_prixunit7.get())
    value7 += 1
    e_q7.delete(0, 'end')
    e_q7.insert(0, value7)
    e_prixtotal7.delete(0, 'end')
    e_prixtotal7.insert(0, str(float(value7)*value27)+ "€")

moins7 = Button(win2, text = "-", command = less7, borderwidth=0)
moins7.grid(row = 9, column = 9)
moins7.config(font = font5, width = 2)
plus7 = Button(win2, text = "+", command = more7, borderwidth=0)
plus7.grid(row = 9, column = 11)
plus7.config(font = font5, width = 2)

quantity7 = StringVar()
e_q7 = Entry(win2, width = 2, fg ="grey", textvariable = quantity7)
e_q7.grid(row = 9, column = 10, sticky = W)
e_q7.config(font = font4)
e_q7.insert(0, "0")

prixtotal7 = StringVar()
e_prixtotal7 = Entry(win2, width = 8, fg ="grey", textvariable = prixtotal7)
e_prixtotal7.grid(row = 9, column = 13, sticky = W)
e_prixtotal7.config(font = font4)


e_designation8 = Entry(win2, width = 30, fg ="grey")
e_designation8.grid(row = 10, column = 2, columnspan = 2, sticky = W)
e_designation8.config(font = font4)

e_u8 = ttk.Combobox(win2, values = unit, width = 7)
e_u8.grid(row = 10, column = 5)

e_prixunit8 = Entry(win2, width = 8, fg ="grey")
e_prixunit8.grid(row = 10, column = 7, sticky = W)
e_prixunit8.config(font = font4)

def less8():
    value8 = int(e_q8.get())
    value28 = float(e_prixunit8.get())
    value8 -= 1
    e_q8.delete(0, 'end')
    e_q8.insert(0, value8)
    e_prixtotal8.delete(0, 'end')
    e_prixtotal8.insert(0, str(float(value8)*value28)+ "€")
def more8():
    value8= int(e_q8.get())
    value28 = float(e_prixunit8.get())
    value8 += 1
    e_q8.delete(0, 'end')
    e_q8.insert(0, value8)
    e_prixtotal8.delete(0, 'end')
    e_prixtotal8.insert(0, str(float(value8)*value28)+ "€")

moins8 = Button(win2, text = "-", command = less8, borderwidth=0)
moins8.grid(row = 10, column = 9)
moins8.config(font = font5, width = 2)
plus8 = Button(win2, text = "+", command = more8, borderwidth=0)
plus8.grid(row = 10, column = 11)
plus8.config(font = font5, width = 2)

quantity8 = StringVar()
e_q8 = Entry(win2, width = 2, fg ="grey", textvariable = quantity8)
e_q8.grid(row = 10, column = 10, sticky = W)
e_q8.config(font = font4)
e_q8.insert(0, "0")

prixtotal8 = StringVar()
e_prixtotal8 = Entry(win2, width = 8, fg ="grey", textvariable = prixtotal8)
e_prixtotal8.grid(row = 10, column = 13, sticky = W)
e_prixtotal8.config(font = font4)



e_designation9 = Entry(win2, width = 30, fg ="grey")
e_designation9.grid(row = 11, column = 2, columnspan = 2, sticky = W)
e_designation9.config(font = font4)

e_u9 = ttk.Combobox(win2, values = unit, width = 7)
e_u9.grid(row = 11, column = 5)

e_prixunit9 = Entry(win2, width = 8, fg ="grey")
e_prixunit9.grid(row = 11, column = 7, sticky = W)
e_prixunit9.config(font = font4)

def less9():
    value9 = int(e_q9.get())
    value29 = float(e_prixunit9.get())
    value9 -= 1
    e_q9.delete(0, 'end')
    e_q9.insert(0, value9)
    e_prixtotal9.delete(0, 'end')
    e_prixtotal9.insert(0, str(float(value9)*value29)+ "€")
def more9():
    value9= int(e_q9.get())
    value29 = float(e_prixunit9.get())
    value9 += 1
    e_q9.delete(0, 'end')
    e_q9.insert(0, value9)
    e_prixtotal9.delete(0, 'end')
    e_prixtotal9.insert(0, str(float(value9)*value29)+ "€")

moins9 = Button(win2, text = "-", command = less9, borderwidth=0)
moins9.grid(row = 11, column = 9)
moins9.config(font = font5, width = 2)
plus9 = Button(win2, text = "+", command = more9, borderwidth=0)
plus9.grid(row = 11, column = 11)
plus9.config(font = font5, width = 2)

quantity9 = StringVar()
e_q9 = Entry(win2, width = 2, fg ="grey", textvariable = quantity9)
e_q9.grid(row = 11, column = 10, sticky = W)
e_q9.config(font = font4)
e_q9.insert(0, "0")

prixtotal9 = StringVar()
e_prixtotal9= Entry(win2, width = 8, fg ="grey", textvariable = prixtotal9)
e_prixtotal9.grid(row = 11, column = 13, sticky = W)
e_prixtotal9.config(font = font4)


e_designation10 = Entry(win2, width = 30, fg ="grey")
e_designation10.grid(row = 12, column = 2, columnspan = 2, sticky = W)
e_designation10.config(font = font4)

e_u10 = ttk.Combobox(win2, values = unit, width = 7)
e_u10.grid(row = 12, column = 5)

e_prixunit10 = Entry(win2, width = 8, fg ="grey")
e_prixunit10.grid(row = 12, column = 7, sticky = W)
e_prixunit10.config(font = font4)

def less10():
    value10 = int(e_q10.get())
    value210 = float(e_prixunit10.get())
    value10 -= 1
    e_q10.delete(0, 'end')
    e_q10.insert(0, value10)
    e_prixtotal10.delete(0, 'end')
    e_prixtotal10.insert(0, str(float(value10)*value210)+ "€")
def more10():
    value10= int(e_q10.get())
    value210 = float(e_prixunit10.get())
    value10 += 1
    e_q10.delete(0, 'end')
    e_q10.insert(0, value10)
    e_prixtotal10.delete(0, 'end')
    e_prixtotal10.insert(0, str(float(value10)*value210)+ "€")

moins10 = Button(win2, text = "-", command = less10, borderwidth=0)
moins10.grid(row = 12, column = 9)
moins10.config(font = font5, width = 2)
plus10 = Button(win2, text = "+", command = more10, borderwidth=0)
plus10.grid(row = 12, column = 11)
plus10.config(font = font5, width = 2)

quantity10 = StringVar()
e_q10 = Entry(win2, width = 2, fg ="grey", textvariable = quantity10)
e_q10.grid(row = 12, column = 10, sticky = W)
e_q10.config(font = font4)
e_q10.insert(0, "0")

prixtotal10 = StringVar()
e_prixtotal10= Entry(win2, width = 8, fg ="grey", textvariable = prixtotal10)
e_prixtotal10.grid(row = 12, column = 13, sticky = W)
e_prixtotal10.config(font = font4)

e_designation11 = Entry(win2, width = 30, fg ="grey")
e_designation11.grid(row = 13, column = 2, columnspan = 2, sticky = W)
e_designation11.config(font = font4)

e_u11 = ttk.Combobox(win2, values = unit, width = 7)
e_u11.grid(row = 13, column = 5)

e_prixunit11 = Entry(win2, width = 8, fg ="grey")
e_prixunit11.grid(row = 13, column = 7, sticky = W)
e_prixunit11.config(font = font4)

def less11():
    value11 = int(e_q11.get())
    value211 = float(e_prixunit11.get())
    value11 -= 1
    e_q11.delete(0, 'end')
    e_q11.insert(0, value11)
    e_prixtotal11.delete(0, 'end')
    e_prixtotal11.insert(0, str(float(value11)*value211)+ "€")
def more11():
    value11= int(e_q11.get())
    value211 = float(e_prixunit11.get())
    value11 += 1
    e_q11.delete(0, 'end')
    e_q11.insert(0, value11)
    e_prixtotal11.delete(0, 'end')
    e_prixtotal11.insert(0, str(float(value11)*value211)+ "€")

moins11 = Button(win2, text = "-", command = less11, borderwidth=0)
moins11.grid(row = 13, column = 9)
moins11.config(font = font5, width = 2)
plus11 = Button(win2, text = "+", command = more11, borderwidth=0)
plus11.grid(row = 13, column = 11)
plus11.config(font = font5, width = 2)

quantity11 = StringVar()
e_q11 = Entry(win2, width = 2, fg ="grey", textvariable = quantity11)
e_q11.grid(row = 13, column = 10, sticky = W)
e_q11.config(font = font4)
e_q11.insert(0, "0")

prixtotal11 = StringVar()
e_prixtotal11= Entry(win2, width = 8, fg ="grey", textvariable = prixtotal11)
e_prixtotal11.grid(row = 13, column = 13, sticky = W)
e_prixtotal11.config(font = font4)


e_designation12 = Entry(win2, width = 30, fg ="grey")
e_designation12.grid(row = 14, column = 2, columnspan = 2, sticky = W)
e_designation12.config(font = font4)

e_u12 = ttk.Combobox(win2, values = unit, width = 7)
e_u12.grid(row = 14, column = 5)

e_prixunit12 = Entry(win2, width = 8, fg ="grey")
e_prixunit12.grid(row = 14, column = 7, sticky = W)
e_prixunit12.config(font = font4)

def less12():
    value12 = int(e_q12.get())
    value212 = float(e_prixunit12.get())
    value12 -= 1
    e_q12.delete(0, 'end')
    e_q12.insert(0, value12)
    e_prixtotal12.delete(0, 'end')
    e_prixtotal12.insert(0, str(float(value12)*value212)+ "€")
def more12():
    value12= int(e_q12.get())
    value212 = float(e_prixunit12.get())
    value12 += 1
    e_q12.delete(0, 'end')
    e_q12.insert(0, value12)
    e_prixtotal12.delete(0, 'end')
    e_prixtotal12.insert(0, str(float(value12)*value212)+ "€")

moins12 = Button(win2, text = "-", command = less12, borderwidth=0)
moins12.grid(row = 14, column = 9)
moins12.config(font = font5, width = 2)
plus12 = Button(win2, text = "+", command = more12, borderwidth=0)
plus12.grid(row = 14, column = 11)
plus12.config(font = font5, width = 2)

quantity12 = StringVar()
e_q12 = Entry(win2, width = 2, fg ="grey", textvariable = quantity12)
e_q12.grid(row = 14, column = 10, sticky = W)
e_q12.config(font = font4)
e_q12.insert(0, "0")

prixtotal12 = StringVar()
e_prixtotal12= Entry(win2, width = 8, fg ="grey", textvariable = prixtotal12)
e_prixtotal12.grid(row = 14, column = 13, sticky = W)
e_prixtotal12.config(font = font4)

e_designation13 = Entry(win2, width = 30, fg ="grey")
e_designation13.grid(row = 15, column = 2, columnspan = 2, sticky = W)
e_designation13.config(font = font4)

e_u13 = ttk.Combobox(win2, values = unit, width = 7)
e_u13.grid(row = 15, column = 5)

e_prixunit13 = Entry(win2, width = 8, fg ="grey")
e_prixunit13.grid(row = 15, column = 7, sticky = W)
e_prixunit13.config(font = font4)

def less13():
    value13= int(e_q13.get())
    value213 = float(e_prixunit13.get())
    value13 -= 1
    e_q13.delete(0, 'end')
    e_q13.insert(0, value13)
    e_prixtotal13.delete(0, 'end')
    e_prixtotal13.insert(0, str(float(value13)*value213)+ "€")
def more13():
    value13= int(e_q13.get())
    value213 = float(e_prixunit13.get())
    value13 += 1
    e_q13.delete(0, 'end')
    e_q13.insert(0, value13)
    e_prixtotal13.delete(0, 'end')
    e_prixtotal13.insert(0, str(float(value13)*value213)+ "€")

moins13 = Button(win2, text = "-", command = less13, borderwidth=0)
moins13.grid(row = 15, column = 9)
moins13.config(font = font5, width = 2)
plus13 = Button(win2, text = "+", command = more13, borderwidth=0)
plus13.grid(row = 15, column = 11)
plus13.config(font = font5, width = 2)

quantity13 = StringVar()
e_q13= Entry(win2, width = 2, fg ="grey", textvariable = quantity13)
e_q13.grid(row = 15, column = 10, sticky = W)
e_q13.config(font = font4)
e_q13.insert(0, "0")

prixtotal13 = StringVar()
e_prixtotal13= Entry(win2, width = 8, fg ="grey", textvariable = prixtotal13)
e_prixtotal13.grid(row = 15, column = 13, sticky = W)
e_prixtotal13.config(font = font4)

e_designation14 = Entry(win2, width = 30, fg ="grey")
e_designation14.grid(row = 16, column = 2, columnspan = 2, sticky = W)
e_designation14.config(font = font4)

e_u14 = ttk.Combobox(win2, values = unit, width = 7)
e_u14.grid(row = 16, column = 5)

e_prixunit14 = Entry(win2, width = 8, fg ="grey")
e_prixunit14.grid(row = 16, column = 7, sticky = W)
e_prixunit14.config(font = font4)

def less14():
    value14 = int(e_q14.get())
    value214= float(e_prixunit14.get())
    value14 -= 1
    e_q14.delete(0, 'end')
    e_q14.insert(0, value14)
    e_prixtotal14.delete(0, 'end')
    e_prixtotal14.insert(0, str(float(value14)*value214)+ "€")
def more14():
    value14= int(e_q14.get())
    value214 = float(e_prixunit14.get())
    value14 += 1
    e_q14.delete(0, 'end')
    e_q14.insert(0, value14)
    e_prixtotal14.delete(0, 'end')
    e_prixtotal14.insert(0, str(float(value14)*value214)+ "€")

moins14 = Button(win2, text = "-", command = less14, borderwidth=0)
moins14.grid(row = 16, column = 9)
moins14.config(font = font5, width = 2)
plus14 = Button(win2, text = "+", command = more14, borderwidth=0)
plus14.grid(row = 16, column = 11)
plus14.config(font = font5, width = 2)

quantity14 = StringVar()
e_q14 = Entry(win2, width = 2, fg ="grey", textvariable = quantity14)
e_q14.grid(row = 16, column = 10, sticky = W)
e_q14.config(font = font4)
e_q14.insert(0, "0")

prixtotal14 = StringVar()
e_prixtotal14= Entry(win2, width = 8, fg ="grey", textvariable = prixtotal14)
e_prixtotal14.grid(row = 16, column = 13, sticky = W)
e_prixtotal14.config(font = font4)

blank5 = Label(win2, text = "(désignation)", borderwidth=8)
blank5.config(font = font4bis, fg="grey")
blank5.grid(row = 17, column = 2, columnspan = 2, sticky = W)

blank51 = Label(win2, text = "(unité)", borderwidth=8)
blank51.config(font = font4bis, fg="grey")
blank51.grid(row = 17, column = 5, columnspan = 2, sticky = W)

blank52 = Label(win2, text = "(€/unit)", borderwidth=8)
blank52.config(font = font4bis, fg="grey")
blank52.grid(row = 17, column = 7, columnspan = 2, sticky = W)

blank53 = Label(win2, text = "(qté)", borderwidth=8, justify='center')
blank53.config(font = font4bis, fg="grey")
blank53.grid(row = 17, column = 9, columnspan=3)

blank54 = Label(win2, text = "(total HT)", borderwidth=8)
blank54.config(font = font4bis, fg="grey")
blank54.grid(row = 17, column = 13, sticky = W)


#PANEL FIN

def effacer():
    messagebox.showinfo("Info", "Pour refaire un bon de commande: \n\nCliquez sur 'OK' puis ouvrez de nouveau le logiciel.")
    
    sys.exit()


blank7 = Label(win1, text = " ", borderwidth=4)
blank7.grid(row = 15, column = 2)

clear = Button(win1, text = "Refaire un BDC", command = effacer, fg = 'white', width = 15)
clear.grid(row = 16, column = 1, columnspan = 2, sticky = W+E)
clear.config(font = font7, bg = '#010083')

blank72 = Label(win1, text = " ", borderwidth=4)
blank72.grid(row = 17, column = 2)

def enregistrer():
    lcg = []
    temp_designation = e_designation.get()
    if (len(temp_designation) != 0):
        temp_u, temp_prixunit, temp_q = e_u.get(), float(e_prixunit.get()), int(e_q.get())
        ligne_cmd = [temp_designation, temp_u, temp_q, temp_prixunit]
        lcg.append(ligne_cmd)
    temp_designation1 = e_designation1.get()
    if (len(temp_designation1) != 0):
        temp_u1, temp_prixunit1, temp_q1 = e_u1.get(), float(e_prixunit1.get()), int(e_q1.get())
        ligne_cmd1 = [temp_designation1, temp_u1, temp_q1, temp_prixunit1]
        lcg.append(ligne_cmd1)
    temp_designation2 = e_designation2.get()
    if (len(temp_designation2) != 0):
        temp_u2, temp_prixunit2, temp_q2 = e_u2.get(), float(e_prixunit2.get()), int(e_q2.get())
        ligne_cmd2 = [temp_designation2, temp_u2, temp_q2, temp_prixunit2]
        lcg.append(ligne_cmd2)
    temp_designation3 = e_designation3.get()
    if (len(temp_designation3) != 0):
        temp_u3, temp_prixunit3, temp_q3 = e_u3.get(), float(e_prixunit3.get()), int(e_q3.get())
        ligne_cmd3 = [temp_designation3, temp_u3, temp_q3, temp_prixunit3]
        lcg.append(ligne_cmd3)
    temp_designation4 = e_designation4.get()
    if (len(temp_designation4) != 0):
        temp_u4, temp_prixunit4, temp_q4 = e_u4.get(), float(e_prixunit4.get()), int(e_q4.get())
        ligne_cmd4 = [temp_designation4, temp_u4, temp_q4, temp_prixunit4]
        lcg.append(ligne_cmd4)
    temp_designation5 = e_designation5.get()
    if (len(temp_designation5) != 0):
        temp_u5, temp_prixunit5, temp_q5 = e_u5.get(), float(e_prixunit5.get()), int(e_q5.get())
        ligne_cmd5 = [temp_designation5, temp_u5, temp_q5, temp_prixunit5]
        lcg.append(ligne_cmd5)
    temp_designation6 = e_designation6.get()
    if (len(temp_designation6) != 0):
        temp_u6, temp_prixunit6, temp_q6 = e_u6.get(), float(e_prixunit6.get()), int(e_q6.get())
        ligne_cmd6 = [temp_designation6, temp_u6, temp_q6, temp_prixunit6]
        lcg.append(ligne_cmd6)
    temp_designation7 = e_designation7.get()
    if (len(temp_designation7) != 0):
        temp_u7, temp_prixunit7, temp_q7 = e_u7.get(), float(e_prixunit7.get()), int(e_q7.get())
        ligne_cmd7 = [temp_designation7, temp_u7, temp_q7, temp_prixunit7]
        lcg.append(ligne_cmd7)
    temp_designation8 = e_designation8.get()
    if (len(temp_designation8) != 0):
        temp_u8, temp_prixunit8, temp_q8 = e_u8.get(), float(e_prixunit8.get()), int(e_q8.get())
        ligne_cmd8 = [temp_designation8, temp_u8, temp_q8, temp_prixunit8]
        lcg.append(ligne_cmd8)
    temp_designation9 = e_designation9.get()
    if (len(temp_designation9) != 0):
        temp_u9, temp_prixunit9, temp_q9 = e_u9.get(), float(e_prixunit9.get()), int(e_q9.get())
        ligne_cmd9 = [temp_designation9, temp_u9, temp_q9, temp_prixunit9]
        lcg.append(ligne_cmd9)
    temp_designation10 = e_designation10.get()
    if (len(temp_designation10) != 0):
        temp_u10, temp_prixunit10, temp_q10 = e_u10.get(), float(e_prixunit10.get()), int(e_q10.get())
        ligne_cmd10 = [temp_designation10, temp_u10, temp_q10, temp_prixunit10]
        lcg.append(ligne_cmd10)
    temp_designation11 = e_designation11.get()
    if (len(temp_designation11) != 0):
        temp_u11, temp_prixunit11, temp_q11 = e_u11.get(), float(e_prixunit11.get()), int(e_q11.get())
        ligne_cmd11 = [temp_designation11, temp_u11, temp_q11, temp_prixunit11]
        lcg.append(ligne_cmd11)
    temp_designation12 = e_designation12.get()
    if (len(temp_designation12) != 0):
        temp_u12, temp_prixunit12, temp_q12 = e_u12.get(), float(e_prixunit12.get()), int(e_q12.get())
        ligne_cmd12 = [temp_designation12, temp_u12, temp_q12, temp_prixunit12]
        lcg.append(ligne_cmd12)
    temp_designation13 = e_designation13.get()
    if (len(temp_designation13) != 0):
        temp_u13, temp_prixunit13, temp_q13 = e_u13.get(), float(e_prixunit13.get()), int(e_q13.get())
        ligne_cmd13 = [temp_designation13, temp_u13, temp_q13, temp_prixunit13]
        lcg.append(ligne_cmd13)
    temp_designation14 = e_designation14.get()
    if (len(temp_designation14) != 0):
        temp_u14, temp_prixunit14, temp_q14 = e_u14.get(), float(e_prixunit14.get()), int(e_q14.get())
        ligne_cmd14 = [temp_designation14, temp_u14, temp_q14, temp_prixunit14]
        lcg.append(ligne_cmd14)


    numero_bdc = str(e_redac.get().rsplit(" ", 1)[0][0]+e_redac.get().rsplit(" ", 1)[1][0])+'-'+str(int(sheetnames_bdc[-3].rsplit('-', 1)[1])+1)
    sheet["B50"] = str(numero_devis.get())
    sheet["C17"] = e_redac.get()
    sheet["C18"] = numero_bdc
    sheet["C20"] = str(e_chantier.get())  #voir si meilleure alternative
    sheet["C16"] = date 
    sheet["C19"] = str((cal.get_date()).strftime("%d/%m/%Y"))
    sheet["E9"] = str(e_chantier.get())
    sheet["E10"] = str(e_adresse.get())
    sheet["E11"] = str(e_cp.get())
    sheet["G12"] = str(e_contact.get()).rsplit(' : ', 1)[0] #personne
    sheet["G13"] = str(e_contact.get()).rsplit(' : ', 1)[1] #numero
    sheet["E16"] = str(e_nom.get())
    sheet["E17"] = str(e_adresse2.get())
    sheet["E18"] = str(e_cp2.get())
    sheet["G19"] = str(e_contact2.get()).rsplit(' : ', 1)[0] #personne
    sheet["G20"] = str(e_contact2.get()).rsplit(' : ', 1)[1] #numero
    if len(str(precisions.get("1.0","end"))) != 0:
        sheet["D22"] = str(precisions.get("1.0","end"))
    else:
        sheet["D22"] = "Aucune précision mentionnée"
    
    print(lcg)
    for j in range(15):
        for c in 'ADFG':
            sheet[f'{c}{26+j}'] = None #flush 
    
    for j in range(len(lcg)):
        if len(lcg[j][0]) != 0:
            i = 0
            for c in 'ADFG':
                sheet[f'{c}{26+j}'] = lcg[j][i]
                i+=1

    current_sheetname = xfile[str(sheetname_bdc2[0])]
    current_sheetname.title = numero_bdc
    xfile.save('ne_pas_toucher.xlsx')

    messagebox.showinfo("Info", "Enregistrement au format XLSX : \n\nCliquez sur 'OK' puis patientez 5 secondes.")

    #UPDATE of DATABASE
    if e_chantier.get() not in livraison["chantier"]:
        with open(pathfiles+"\\Lchantier.txt", "a", encoding = 'utf-8') as file:
            file.write('\n')
            file.write(e_chantier.get())
        with open(pathfiles+"\\Ladresse.txt", "a", encoding = 'utf-8') as file:
            file.write('\n')
            file.write(e_adresse.get())
        with open(pathfiles+"\\Lville.txt", "a", encoding = 'utf-8') as file:
            file.write('\n')
            file.write(e_cp.get())
        with open(pathfiles+"\\Lnumero.txt", "a", encoding = 'utf-8') as file:
            file.write('\n')
            file.write(''.join(re.findall('\d',e_contact.get())))
        with open(pathfiles+"\\Lcontact.txt", "a", encoding = 'utf-8') as file:
            file.write('\n')
            file.write(''.join(re.sub(r'[^a-zA-Z]', ' ', e_contact.get())))
        
    if e_nom.get() not in fournisseur["nom"]:
        with open("Fentreprise.txt", "a", encoding = 'utf-8') as file:
            file.write('\n')
            file.write(e_nom.get())
        with open("Fadresse.txt", "a", encoding = 'utf-8') as file:
            file.write('\n')
            file.write(e_adresse2.get())
        with open("Fville.txt", "a", encoding = 'utf-8') as file:
            file.write('\n')
            file.write(e_cp2.get())
        with open("Fnumero.txt", "a", encoding = 'utf-8') as file:
            file.write('\n')
            file.write(''.join(re.findall('\d',e_contact2.get())))
        with open("Fcontact.txt", "a", encoding = 'utf-8') as file:
            file.write('\n')
            file.write(''.join(re.sub(r'[^a-zA-Z]', ' ', e_contact2.get())))

    def copytobdc():
        path1 = os.getcwd()+'\\ne_pas_toucher.xlsx'
        path2 = os.getcwd()+'\\BDC - EQUO CONSTRUCTION.xlsx'

        xl = Dispatch("Excel.Application")
        xl.Visible = False 

        wb1 = xl.Workbooks.Open(Filename=path1)
        wb2 = xl.Workbooks.Open(Filename=path2)

        ws1 = wb1.Worksheets(1)
        ws1.Copy(Before = wb2.Worksheets("SUIVI"))
        wb2.Close(SaveChanges=True)
        wb1.Close(SaveChanges=True)
        xl.Quit()
        
        window.wm_state('iconic')
        window.wm_state('normal')
    
    copytobdc()

sauvegarder = Button(win1, text = "Enregistrer (.xlsx)", command = enregistrer, fg = 'white', width = 24)
sauvegarder.grid(row = 14, column = 1, sticky = W)
sauvegarder.config(font = font7, bg = '#1f6e43')

def imprimer():
    # print(str(cal.get_date()))
    # print((cal.get_date()).strftime("%d/%m/%Y"))
    numero_bdc = str(e_redac.get().rsplit(" ", 1)[0][0]+e_redac.get().rsplit(" ", 1)[1][0])+'-'+str(int(sheetnames_bdc[-3].rsplit('-', 1)[1])+1)
    
    WB_PATH = os.getcwd()+r'\\ne_pas_toucher.xlsx'
    PATH_TO_PDF = os.getcwd()+r'\\'+numero_bdc+'.pdf'

    excel = win32com.client.Dispatch("Excel.Application")

    try:
        wb = excel.Workbooks.Open(WB_PATH)
        ws_index_list = [1]
        wb.WorkSheets(ws_index_list).Select()
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
    except com_error as e:
        print('échec')
    else:
        print('réussi')
    wb.Close()
    excel.Quit()

    messagebox.showinfo("Info", "Enregistrement au format PDF : succès.\n\nVeuillez fermer l'application après impression.")
    subprocess.Popen([+numero_bdc+'.pdf'],shell=True)


topdf = Button(win1, text = "Imprimer (.pdf)", command = imprimer, fg = 'white', width = 24)
topdf.grid(row = 14, column = 2, sticky = E)
topdf.config(font = font7, bg = '#aa0a00')




window.mainloop()